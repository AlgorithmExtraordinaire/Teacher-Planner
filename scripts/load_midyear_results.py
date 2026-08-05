#!/usr/bin/env python3
"""
Load the SCA First-Semester (Mid-Year) Results workbook into Supabase.

    cd /opt/teacher-planner
    SUPA_URL=$(grep '^NEXT_PUBLIC_SUPABASE_URL=' .env | cut -d= -f2-) \
    SUPA_KEY=$(grep '^SUPABASE_SERVICE_ROLE_KEY=' .env | cut -d= -f2-) \
    python3 scripts/load_midyear_results.py /path/to/results.xlsx

Reads four sheets — Mid-Year Results, Non-Promotional, Attendance, Comments —
and writes assessments, assessment_results, attendance_summary and
report_comments.

DESIGN NOTES
------------
Learners are matched on Roll No. against students.student_number, never on
name. The workbook writes names surname-first ("Cassim Surei") while the
roster holds them given-name-first, so name matching would mis-link learners
— the worst possible failure in an assessment system.

sbg_level is deliberately left NULL. The workbook records percentages; the
school's Standards-Based Grading scale is 1-4, and no published mapping
between the two exists in the Knowledge Base. Inventing thresholds would
manufacture grades nobody awarded.

Promotional subjects link to the real class for that grade. Non-promotional
subjects (Character Education, PE, Arts, Computer Tech, Foreign Language)
have no class in the roster, so their assessments are recorded with a null
class_id rather than being attached to an unrelated class.

Idempotent throughout: assessments upsert on (title, class_id), results on
(assessment_id, student_id), attendance and comments on (student_id, term).
"""

import json
import os
import sys
import urllib.parse
import urllib.request

TERM = "Mid-Year 2026"

PROMOTIONAL = {
    "Maths": "Mathematics",
    "English Language Arts": "English Language Arts",
    "Science": "Science",
    "Social Studies": "Social Studies",
}

NON_PROMOTIONAL = [
    "Character Education", "Physical Education", "Arts",
    "Computer Tech", "Foreign Language (Duolingo)", "Art", "Music",
]

URL = os.environ["SUPA_URL"].rstrip("/")
KEY = os.environ["SUPA_KEY"]
HEADERS = {
    "apikey": KEY,
    "Authorization": f"Bearer {KEY}",
    "Content-Type": "application/json",
}


def get(path):
    req = urllib.request.Request(f"{URL}/rest/v1/{path}", headers=HEADERS)
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read())


def upsert(table, rows, conflict, chunk=200):
    if not rows:
        return 0
    hdr = dict(HEADERS)
    hdr["Prefer"] = "resolution=merge-duplicates,return=minimal"
    sent = 0
    for i in range(0, len(rows), chunk):
        part = rows[i:i + chunk]
        q = urllib.parse.urlencode({"on_conflict": conflict})
        req = urllib.request.Request(
            f"{URL}/rest/v1/{table}?{q}",
            data=json.dumps(part).encode(), headers=hdr, method="POST",
        )
        with urllib.request.urlopen(req, timeout=90):
            sent += len(part)
    return sent


def num(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)

    students = {s["student_number"]: s["id"]
                for s in get("students?select=id,student_number")
                if s.get("student_number")}
    classes = {(c["grade_level"], c["subject"]): c["id"]
               for c in get("classes?select=id,grade_level,subject")}
    print(f"roster: {len(students)} learners, {len(classes)} classes")

    def sheet_rows(name):
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip() if h else "" for h in rows[0]]
        return header, rows[1:]

    # ---- assessments -------------------------------------------------
    wanted = set()
    hdr, rows = sheet_rows("Mid-Year Results")
    for r in rows:
        if not r or not r[0]:
            continue
        grade = f"Grade {str(r[0]).strip()}"
        for col, subject in PROMOTIONAL.items():
            if col in hdr:
                wanted.add((grade, subject))

    assessments = [{
        "title": f"{TERM} — {subject} ({grade})",
        "class_id": classes.get((grade, subject)),
        "type": "summative",
        "sbg_level_max": 4,
    } for grade, subject in sorted(wanted)]

    for subject in NON_PROMOTIONAL:
        assessments.append({
            "title": f"{TERM} — {subject}",
            "class_id": None, "type": "non_promotional", "sbg_level_max": 4,
        })

    upsert("assessments", assessments, "title,class_id")
    lookup = {a["title"]: a["id"] for a in get("assessments?select=id,title")}
    print(f"assessments: {len(lookup)}")

    # ---- results -----------------------------------------------------
    results, skipped = [], []

    def collect(sheet, columns, titler):
        hdr, rows = sheet_rows(sheet)
        idx = {h: i for i, h in enumerate(hdr)}
        for r in rows:
            if not r or not r[0]:
                continue
            roll = str(r[idx["Roll No."]]).strip()
            sid = students.get(roll)
            if not sid:
                skipped.append(roll)
                continue
            grade = f"Grade {str(r[0]).strip()}"
            for col in columns:
                if col not in idx:
                    continue
                score = num(r[idx[col]])
                if score is None:
                    continue
                aid = lookup.get(titler(col, grade))
                if aid:
                    results.append({"assessment_id": aid, "student_id": sid,
                                    "score": score})

    collect("Mid-Year Results", list(PROMOTIONAL),
            lambda c, g: f"{TERM} — {PROMOTIONAL[c]} ({g})")
    collect("Non-Promotional", NON_PROMOTIONAL,
            lambda c, g: f"{TERM} — {c}")

    print("results:", upsert("assessment_results", results,
                             "assessment_id,student_id"))
    if skipped:
        print(f"  skipped {len(skipped)} rows with no matching roll number: "
              f"{sorted(set(skipped))[:8]}")

    # ---- attendance --------------------------------------------------
    hdr, rows = sheet_rows("Attendance")
    idx = {h: i for i, h in enumerate(hdr)}
    att = []
    for r in rows:
        if not r or not r[0]:
            continue
        sid = students.get(str(r[idx["Roll No."]]).strip())
        if not sid:
            continue
        pct = num(r[idx["Attendance %"]])
        att.append({
            "student_id": sid, "term": TERM,
            "present_days": int(num(r[idx["Present Days"]]) or 0),
            "total_days": int(num(r[idx["Total Days"]]) or 0),
            # stored as a percentage, not the workbook's 0-1 fraction
            "attendance_pct": round(pct * 100, 2) if pct and pct <= 1 else pct,
        })
    print("attendance:", upsert("attendance_summary", att, "student_id,term"))

    # ---- comments ----------------------------------------------------
    hdr, rows = sheet_rows("Comments")
    idx = {h: i for i, h in enumerate(hdr)}
    comments = []
    for r in rows:
        if not r or not r[0]:
            continue
        sid = students.get(str(r[idx["Roll No."]]).strip())
        text = r[idx["Teacher Comment"]]
        if sid and text:
            comments.append({"student_id": sid, "term": TERM,
                             "comment": str(text).strip()})
    print("comments:", upsert("report_comments", comments, "student_id,term"))


if __name__ == "__main__":
    main(sys.argv[1])
